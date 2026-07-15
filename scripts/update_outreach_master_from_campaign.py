#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import shutil
from collections import OrderedDict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


NEW_COLUMNS = [
    "first_email_status",
    "first_email_subject",
    "first_email_body",
    "first_email_sent_at",
    "first_email_source_file",
    "delivery_status",
    "bounce_notice_at",
    "bounce_source",
    "email_needs_review",
    "replacement_email_candidates",
    "second_email_subject",
    "second_email_body",
]


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def director_name(full_name: str) -> str:
    parts = [p for p in normalize_text(full_name).split() if p]
    if len(parts) >= 2:
        return f"{parts[1]} {parts[0]}"
    return normalize_text(full_name)


def activity_hint(activity: str) -> str:
    text = normalize_text(activity).lower()
    mapping = [
        ("транспорт", "рабочих рисков в транспортной и договорной работе"),
        ("груз", "рабочих рисков в транспортной и договорной работе"),
        ("сельск", "изменений и рисков для сельхозпредприятий"),
        ("выращив", "изменений и рисков для сельхозпредприятий"),
        ("строит", "рабочих рисков в договорах и документообороте"),
        ("торгов", "рабочих рисков в договорах, кадрах и налоговых вопросах"),
        ("образован", "рабочих вопросов по кадровой и договорной части"),
        ("медицин", "рабочих вопросов по проверкам и документальному оформлению"),
        ("бухгалтер", "рабочих вопросов по проверкам и документальному оформлению"),
        ("юрид", "рабочих вопросов по проверкам и документальному оформлению"),
    ]
    for needle, result in mapping:
        if needle in text:
            return result
    return "ежедневных рабочих вопросов по законодательству, договорам и документам"


def build_followup_subject(company: str, activity: str) -> str:
    hint = activity_hint(activity)
    if "сельхоз" in hint:
        return "Повторно по вопросам работы сельхозпредприятий"
    if "транспорт" in hint:
        return "Повторно по вопросам логистики и документооборота"
    return "Пишу повторно по КонсультантПлюс"


def build_followup_body(record: dict) -> str:
    name = director_name(record.get("director") or "")
    company = normalize_text(record.get("company"))
    activity = normalize_text(record.get("activity"))
    hint = activity_hint(activity)
    return (
        f"{name}, добрый день!\n\n"
        "Недавно уже направляли вам письмо, поэтому коротко повторю суть.\n\n"
        f"На практике у {company or 'компаний вашего профиля'} больше всего времени обычно уходит на решение {hint}. "
        "Именно в таких задачах КонсультантПлюс помогает быстрее находить ответы, снижать риск ошибок и не тратить лишнее время на поиск позиции по спорным вопросам.\n\n"
        "Если для вас это актуально, можем показать на нескольких типовых ситуациях, как это работает в ежедневной работе.\n\n"
        "С уважением,\n"
        "Иванова Вероника\n"
        "КонсультантПлюс Волгоград\n"
        "8(8442) 43-27-27"
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("master_xlsx", type=Path)
    parser.add_argument("letters_json", type=Path)
    parser.add_argument("--backup", action="store_true")
    parser.add_argument("--followup-start", type=int, default=0)
    parser.add_argument("--followup-count", type=int, default=10)
    args = parser.parse_args()

    if args.backup:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = args.master_xlsx.with_suffix(f".backup_{stamp}.xlsx")
        shutil.copy2(args.master_xlsx, backup_path)

    with args.letters_json.open("r", encoding="utf-8") as fh:
        letters = json.load(fh)

    letters_by_inn = OrderedDict()
    for item in letters:
        inn = normalize_text(item.get("inn"))
        if inn:
            letters_by_inn[inn] = item

    wb = load_workbook(args.master_xlsx)
    ws = wb[wb.sheetnames[0]]

    headers = [cell.value for cell in ws[1]]
    header_index = {normalize_text(name): idx + 1 for idx, name in enumerate(headers)}
    for column in NEW_COLUMNS:
        if column not in header_index:
            ws.cell(row=1, column=ws.max_column + 1, value=column)
            header_index[column] = ws.max_column

    inn_col = header_index["inn"]
    followup_seen = 0
    followup_written = 0
    for row_idx in range(2, ws.max_row + 1):
        inn = normalize_text(ws.cell(row=row_idx, column=inn_col).value)
        if not inn or inn not in letters_by_inn:
            continue
        item = letters_by_inn[inn]
        ws.cell(row=row_idx, column=header_index["first_email_status"], value=item.get("status"))
        ws.cell(row=row_idx, column=header_index["first_email_subject"], value=item.get("subject"))
        ws.cell(row=row_idx, column=header_index["first_email_body"], value=item.get("body"))
        ws.cell(row=row_idx, column=header_index["first_email_sent_at"], value=item.get("sent_at"))
        ws.cell(row=row_idx, column=header_index["first_email_source_file"], value=item.get("_source_file"))

        delivery_status = item.get("delivery_status") or ("bounced" if item.get("status") == "bounced" else "")
        if delivery_status:
            ws.cell(row=row_idx, column=header_index["delivery_status"], value=delivery_status)
        if item.get("bounce_notice_at"):
            ws.cell(row=row_idx, column=header_index["bounce_notice_at"], value=item.get("bounce_notice_at"))
        if item.get("bounce_source"):
            ws.cell(row=row_idx, column=header_index["bounce_source"], value=item.get("bounce_source"))
        if item.get("email_needs_review") is not None:
            ws.cell(row=row_idx, column=header_index["email_needs_review"], value=str(bool(item.get("email_needs_review"))).lower())
        elif delivery_status == "bounced":
            ws.cell(row=row_idx, column=header_index["email_needs_review"], value="true")
        candidates = item.get("replacement_email_candidates") or []
        if candidates:
            rendered = "; ".join(
                f"{c.get('email')} [{c.get('confidence')}: {c.get('reason')}]"
                for c in candidates
            )
            ws.cell(row=row_idx, column=header_index["replacement_email_candidates"], value=rendered)

        if normalize_text(item.get("status")) == "sent":
            if followup_seen < args.followup_start:
                followup_seen += 1
                continue
            if followup_written >= args.followup_count:
                continue
            ws.cell(
                row=row_idx,
                column=header_index["second_email_subject"],
                value=build_followup_subject(item.get("company") or "", item.get("activity") or ""),
            )
            ws.cell(
                row=row_idx,
                column=header_index["second_email_body"],
                value=build_followup_body(item),
            )
            followup_seen += 1
            followup_written += 1

    wb.save(args.master_xlsx)


if __name__ == "__main__":
    main()
